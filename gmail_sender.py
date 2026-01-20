import os
import base64
import pickle
from io import BytesIO
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image, ImageDraw, ImageFont

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build


# =========================
# CONFIGURATION
# =========================
SENDER_EMAIL = "anjali.rathore@coolbootsmedia.co"
RECIPIENT_EMAIL = os.getenv('RECIPIENT_EMAIL', "ankit.k@coolbootsmedia.com")
EMAIL_SUBJECT = "📊 Automated Report - Registration Template"
EMAIL_BODY = "📊 Here is the Automated Report 4\n\nPlease find the registration template attached."
EXCEL_FILENAME = "Registration_Template.xlsx"
SHEET_NAME = "Template"

# Gmail API scope
SCOPES = ['https://www.googleapis.com/auth/gmail.send']


# =========================
# CROSS-PLATFORM EXCEL TO IMAGE (IMPROVED)
# =========================
def hex_to_rgb(hex_color):
    """Convert hex color to RGB tuple"""
    if hex_color.startswith('#'):
        hex_color = hex_color[1:]
    if len(hex_color) == 8:  # ARGB format
        hex_color = hex_color[2:]  # Remove alpha channel
    if len(hex_color) == 6:
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    return (255, 255, 255)  # Default to white


def excel_to_image_cross_platform(excel_path, sheet_name):
    """
    Convert Excel sheet to PNG image with proper styling
    Improved version with better color handling and clarity
    """
    print(f"📂 Opening Excel file: {excel_path}")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    try:
        ws = wb[sheet_name]
        print(f"✅ Opened sheet: '{sheet_name}'")
    except KeyError:
        print(f"⚠️ Sheet '{sheet_name}' not found. Using first sheet.")
        ws = wb.active
    
    # Get used range
    max_row = ws.max_row
    max_col = ws.max_column
    
    print(f"📊 Processing {max_row} rows × {max_col} columns")
    
    # Read data and styles
    data = []
    styles = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col), 1):
        row_data = []
        row_styles = []
        
        for col_idx, cell in enumerate(row, 1):
            # Get cell value
            value = str(cell.value) if cell.value is not None else ''
            row_data.append(value)
            
            # Get cell style
            cell_style = {
                'bold': cell.font.bold if cell.font else False,
                'bg_color': (255, 255, 255)  # Default white background
            }
            
            # Get background color
            if cell.fill and cell.fill.start_color:
                if hasattr(cell.fill.start_color, 'rgb'):
                    rgb_value = cell.fill.start_color.rgb
                    if rgb_value and rgb_value != '00000000':
                        cell_style['bg_color'] = hex_to_rgb(rgb_value)
            
            row_styles.append(cell_style)
        
        data.append(row_data)
        styles.append(row_styles)
    
    # Calculate image dimensions (larger for better clarity)
    cell_width = 150  # Increased from 120
    cell_height = 40  # Increased from 35
    padding = 15
    
    # Calculate column widths based on content
    col_widths = []
    for col_idx in range(max_col):
        max_length = 0
        for row_data in data:
            if col_idx < len(row_data):
                max_length = max(max_length, len(str(row_data[col_idx])))
        
        # First column (Source) wider, others based on content
        if col_idx == 0:
            col_widths.append(max(200, max_length * 10))
        else:
            col_widths.append(max(120, max_length * 10))
    
    img_width = sum(col_widths) + padding * 2
    img_height = max_row * cell_height + padding * 2
    
    print(f"🖼️ Creating image: {img_width}x{img_height} pixels")
    
    # Create image with white background
    img = Image.new('RGB', (img_width, img_height), color='white')
    draw = ImageDraw.Draw(img)
    
    # Try to use better fonts
    try:
        # Try different font paths for different systems
        font_regular = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 14)
        font_bold = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 14)
    except:
        try:
            font_regular = ImageFont.truetype("arial.ttf", 14)
            font_bold = ImageFont.truetype("arialbd.ttf", 14)
        except:
            print("⚠️ Using default font (may have lower quality)")
            font_regular = ImageFont.load_default()
            font_bold = font_regular
    
    # Draw cells
    y_offset = padding
    for row_idx, (row_data, row_style) in enumerate(zip(data, styles)):
        x_offset = padding
        
        for col_idx, (cell_value, cell_style) in enumerate(zip(row_data, row_style)):
            col_width = col_widths[col_idx]
            
            # Draw cell background
            bg_color = cell_style['bg_color']
            draw.rectangle(
                [x_offset, y_offset, x_offset + col_width, y_offset + cell_height],
                fill=bg_color,
                outline='black',
                width=1
            )
            
            # Choose font (bold or regular)
            current_font = font_bold if cell_style['bold'] else font_regular
            
            # Calculate text position (centered)
            try:
                bbox = draw.textbbox((0, 0), cell_value, font=current_font)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
            except:
                # Fallback for older PIL versions
                text_width = len(cell_value) * 8
                text_height = 14
            
            # Center text horizontally and vertically
            text_x = x_offset + (col_width - text_width) / 2
            text_y = y_offset + (cell_height - text_height) / 2
            
            # Draw text in black
            draw.text((text_x, text_y), cell_value, fill='black', font=current_font)
            
            x_offset += col_width
        
        y_offset += cell_height
    
    wb.close()
    print(f"✅ Excel sheet '{sheet_name}' converted to image successfully")
    
    return img


# =========================
# GMAIL AUTHENTICATION
# =========================
def authenticate_gmail():
    """Authenticate with Gmail API using OAuth 2.0"""
    creds = None
    
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            print("🔄 Refreshing authentication token...")
            creds.refresh(Request())
        else:
            print("🔐 First-time authentication required...")
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            # Use run_console for headless environments
            creds = flow.run_console() if os.getenv('CI') else flow.run_local_server(port=0)
        
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)
        print("✅ Authentication successful!")
    else:
        print("✅ Using saved authentication")
    
    service = build('gmail', 'v1', credentials=creds)
    return service


# =========================
# EMAIL CREATION & SENDING
# =========================
def create_message_with_attachment(sender, to, subject, body_text, img):
    """Create email message with image attachment"""
    message = MIMEMultipart()
    message['to'] = to
    message['from'] = sender
    message['subject'] = subject

    msg_body = MIMEText(body_text, 'plain')
    message.attach(msg_body)

    # Convert PIL Image to bytes with high quality
    img_byte_arr = BytesIO()
    img.save(img_byte_arr, format='PNG', optimize=False, quality=100)
    img_byte_arr = img_byte_arr.getvalue()

    image_attachment = MIMEImage(img_byte_arr, name='registration_template.png')
    message.attach(image_attachment)

    raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')
    return {'raw': raw_message}


def send_email(service, message):
    """Send email using Gmail API"""
    try:
        sent_message = service.users().messages().send(
            userId='me', body=message).execute()
        print(f"✅ Email sent successfully! Message ID: {sent_message['id']}")
        return sent_message
    except Exception as e:
        print(f"❌ Error sending email: {e}")
        raise


# =========================
# MAIN FUNCTION
# =========================
def send_registration_template_via_gmail(
    recipient=None,
    subject=EMAIL_SUBJECT,
    body=EMAIL_BODY
):
    """Main function to send Registration Template via Gmail"""
    print("\n📧 Starting Gmail automation...\n")
    
    recipient = recipient or RECIPIENT_EMAIL
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, EXCEL_FILENAME)
    
    credentials_path = os.path.join(script_dir, 'credentials.json')
    if not os.path.exists(credentials_path):
        raise FileNotFoundError(
            f"❌ credentials.json not found!\n"
            f"   Please download it from Google Cloud Console"
        )
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"❌ Excel file not found: {excel_path}")
    
    print(f"📂 Found Excel file: {excel_path}\n")
    
    # Step 1: Convert Excel to Image
    print("STEP 1: Converting Excel to Image...")
    img = excel_to_image_cross_platform(excel_path, SHEET_NAME)
    
    # Step 2: Authenticate with Gmail
    print("\nSTEP 2: Authenticating with Gmail API...")
    service = authenticate_gmail()
    
    # Step 3: Create email message
    print("\nSTEP 3: Creating email message...")
    message = create_message_with_attachment(
        sender=SENDER_EMAIL,
        to=recipient,
        subject=subject,
        body_text=body,
        img=img
    )
    print(f"   To: {recipient}")
    print(f"   Subject: {subject}")
    
    # Step 4: Send email
    print("\nSTEP 4: Sending email...")
    send_email(service, message)
    
    print("\n🎉 Gmail automation complete!\n")


if __name__ == "__main__":
    send_registration_template_via_gmail()