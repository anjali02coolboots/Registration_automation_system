import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import os

def generate_excel_template(input_file):
    """
    Generate formatted Excel template from registration data and add as new sheet
    
    Parameters:
    input_file (str): Path to the input Excel file (Registration_Template.xlsx)
    """
    
    try:
        # Read the input data
        df = pd.read_excel(input_file)
        
        # Check for duplicate dates
        print("\n[DUPLICATE CHECK] Analyzing data for duplicate dates...")
        df['Date'] = pd.to_datetime(df['Date'], format='%d-%m-%Y', errors='coerce')
        
        # Count rows per date
        date_counts = df['Date'].value_counts().sort_index(ascending=False)
        
        # Check for duplicates
        duplicate_dates = date_counts[date_counts > 1]
        if len(duplicate_dates) > 0:
            print(f"⚠️  WARNING: Found duplicate data for {len(duplicate_dates)} date(s):")
            for date, count in duplicate_dates.items():
                print(f"   - {date.strftime('%d-%m-%Y')}: {count} entries")
            
            print("\n[DEDUPLICATION] This shouldn't happen with the updated data_processor.py")
            print("                If you see this, the duplicate detection in data_processor.py may need attention.")
        else:
            print("✓ No duplicate dates found - data is clean!")
        
        # Get min and max dates
        min_date = df['Date'].min()
        max_date = df['Date'].max()
        
        print(f"\n[DATE RANGE] Processing data from {min_date.strftime('%d-%m-%Y')} to {max_date.strftime('%d-%m-%Y')}")
        print(f"[DATE RANGE] Total unique dates: {df['Date'].nunique()}")
        print(f"[DATE RANGE] Total rows: {len(df)}")
        
        # Generate all dates in the range
        date_range = pd.date_range(start=min_date, end=max_date, freq='D')
        
        # Define the required source categories in specific order
        required_sources = ['content.techgig.com', 'Organic', 'Delivery', 'Social Media']
        
        # Get all unique sources from data
        unique_sources = df['New Source'].unique().tolist()
        
        # Combine required sources with any additional sources found in data
        all_sources = required_sources.copy()
        for source in unique_sources:
            if source not in all_sources and pd.notna(source):
                all_sources.append(source)
        
        # Create output data structure
        output_data = []
        
        # Process each source
        for source in all_sources:
            row = {'Source': source}
            
            # Count registrations for each date
            for date in date_range:
                date_str = date.strftime('%m-%d-%Y')
                
                # Filter data for this source and date
                count = len(df[(df['New Source'] == source) & (df['Date'] == date)])
                
                # Add to row (use '-' for zero counts)
                row[date_str] = count if count > 0 else '-'
            
            # Calculate total for this source
            total = len(df[df['New Source'] == source])
            row['Total'] = total
            
            output_data.append(row)
        
        # Add "Total of Registration" row
        total_row = {'Source': 'Total of Registration'}
        for date in date_range:
            date_str = date.strftime('%m-%d-%Y')
            count = len(df[df['Date'] == date])
            total_row[date_str] = count if count > 0 else '-'
        
        # Overall total
        total_row['Total'] = len(df)
        output_data.append(total_row)
        
        # Create DataFrame
        result_df = pd.DataFrame(output_data)
        
        # Reorder columns: Source, dates in order, Total
        date_columns = [date.strftime('%m-%d-%Y') for date in date_range]
        column_order = ['Source'] + date_columns + ['Total']
        result_df = result_df[column_order]
        
        # Load the existing workbook
        wb = load_workbook(input_file)
        
        # Remove "Template" sheet if it already exists
        if "Template" in wb.sheetnames:
            del wb["Template"]
        
        # Create new "Template" sheet
        ws = wb.create_sheet("Template")
        
        # Write headers
        headers = result_df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Write data rows
        for row_idx, row_data in enumerate(result_df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Format "Total of Registration" row
                if row_data[0] == 'Total of Registration':
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                
                # Align first column (Source) to left, others to center
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25  # Source column
        for col_idx in range(2, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 12
        
        # Save the workbook (overwrites the original file)
        wb.save(input_file)
        
        print("\n" + "="*60)
        print("TEMPLATE GENERATION COMPLETE!")
        print("="*60)
        print(f"✓ Template sheet created: {input_file}")
        print(f"✓ Date range: {min_date.strftime('%m-%d-%Y')} to {max_date.strftime('%m-%d-%Y')}")
        print(f"✓ Total sources: {len(all_sources)}")
        print(f"✓ Total registrations: {len(df)}")
        print(f"✓ Unique dates processed: {df['Date'].nunique()}")
        print("="*60)
        
        return input_file
        
    except FileNotFoundError:
        print(f"✗ Error: File '{input_file}' not found.")
        return None
    except Exception as e:
        print(f"✗ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


# Usage
if __name__ == "__main__":
    input_file = "Registration_Template.xlsx"
    
    # Check if file exists
    if os.path.exists(input_file):
        generate_excel_template(input_file)
    else:
        print(f"✗ Error: File not found at '{input_file}'")
        print(f"Current directory: {os.getcwd()}")
        print("\nPlease ensure the file is in the same directory as the script:")
        print("  Report 4/")
        print("    ├── generate_template.py")
        print("    └── Registration_Template.xlsx")